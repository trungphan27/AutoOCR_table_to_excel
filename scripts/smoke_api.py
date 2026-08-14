"""Exercise health, JSON inference and Excel endpoints in-process."""

import argparse
import io
import os
import sys
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from deploy.envfile import load_env_file


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("image", type=Path)
    parser.add_argument("--env-file", type=Path, default=Path(".env"))
    parser.add_argument("--with-gradio", action="store_true")
    return parser.parse_args()


def main():
    args = parse_args()
    load_env_file(args.env_file)
    if not args.with_gradio:
        os.environ["OCR_ENABLE_GRADIO"] = "false"
    image = args.image.expanduser().resolve()
    if not image.is_file():
        raise FileNotFoundError(image)
    content = image.read_bytes()

    # App feature flags are evaluated during import.
    from deploy.app import app

    with TestClient(app) as client:
        live = client.get("/live")
        ready = client.get("/ready")
        recognize = client.post(
            "/v1/table/recognize",
            files={"file": (image.name, content, "image/png")},
        )
        excel = client.post(
            "/v1/table/excel",
            files={"file": (image.name, content, "image/png")},
        )
    for name, response in {
        "live": live,
        "ready": ready,
        "recognize": recognize,
        "excel": excel,
    }.items():
        if response.status_code != 200:
            raise RuntimeError(
                "{} failed with {}: {}".format(
                    name, response.status_code, response.text
                )
            )
    payload = recognize.json()
    workbook = load_workbook(
        io.BytesIO(excel.content), read_only=True, data_only=True
    )
    sheet_names = workbook.sheetnames
    workbook.close()
    print("ready={}".format(ready.json()))
    print("timing={}".format(payload["timing"]))
    print("html_chars={}".format(len(payload["result"]["html"])))
    print("excel_bytes={} sheets={}".format(len(excel.content), sheet_names))


if __name__ == "__main__":
    main()
