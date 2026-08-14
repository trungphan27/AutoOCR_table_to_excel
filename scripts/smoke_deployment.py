"""Run one end-to-end image -> HTML -> XLSX deployment smoke test."""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from deploy.envfile import load_env_file
from deploy.service import TableInferenceService


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("image", type=Path)
    parser.add_argument("--env-file", type=Path, default=Path(".env"))
    parser.add_argument("--keep-excel", action="store_true")
    return parser.parse_args()


def main():
    args = parse_args()
    load_env_file(args.env_file)
    image = args.image.expanduser().resolve()
    if not image.is_file():
        raise FileNotFoundError(image)
    service = TableInferenceService.from_env()
    payload = service.predict_bytes(image.read_bytes())
    excel = service.save_excel(payload)
    workbook = load_workbook(excel, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    print("providers={}".format(service.provider_info()))
    print("timing={}".format(payload["timing"]))
    print("html_chars={}".format(len(payload["result"]["html"])))
    print("excel={} sheets={}".format(excel, sheet_names))
    if not args.keep_excel:
        excel.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
